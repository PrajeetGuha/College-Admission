import os
from tkinter.constants import TRUE
import openpyxl
import pathlib

def check_file(loc,info=0):
    
    '''
    Checks if a file i present or not. If not present creates the file
    '''
    
    file = pathlib.Path(loc)
    if not(file.exists()):
        if info == 0:
            data=["Email","Password"]
        else:
            data=['Registration Number','Name','DOB','Father\'s Name','Mother\'s Name','Gender','Dept','Domicile','Mother\'s Phone Number','Father\'s Phone Number','Email Id','Present Address','Permanent Address','High School Name','High School Board Name','High School Marks','Higher Secondary School Name','Higher Secondary Board Name', 'High Secondary Marks(P,C,M,(B/Co),Eng)']

        wb =openpyxl.workbook.Workbook()
        ws0 = wb.worksheets[0]
        ws0.append(data)
        wb.save(filename =loc)

def check_email_present(email,log):
    
    '''
    Checks if the email is present in the login file or not
    '''
    
    if log == 1:
        loc=os.path.dirname(__file__) +'\databases\studentlogininfo.xlsx'
    else:
        loc=os.path.dirname(__file__) +'\databases\adminlogininfo.xlsx'
    check_file(loc)
    book =openpyxl.load_workbook(loc)
    sheet = book.active
    for i in range(1,sheet.max_row+1):
        if(((sheet.cell(row=i,column=1)).value)==email):
            return True
    return False

def load(email):
    
    '''
    It loads the saved data from the student information database
    '''
    
    loc2=os.path.dirname(__file__) +'\databases\studentinfo.xlsx'
    check_file(loc2,1)
    book =openpyxl.load_workbook(loc2)
    sheet = book.active
    info_dict={}
    for i in range(1,sheet.max_row+1):
        if(((sheet.cell(row=i,column=11)).value)==email):
            for j in range(1,20):
                if (sheet.cell(row=i,column=j)).value == None:
                    info_dict[(sheet.cell(row=1,column=j)).value]=''
                else:
                    info_dict[(sheet.cell(row=1,column=j)).value]=(sheet.cell(row=i,column=j)).value
            info_dict['record']=i
            print(info_dict)
            return info_dict
        
def creditional_check(email,password,log):
    
    '''
    Checks if the credentials match with the database
    '''
    
    if log == 1:
        loc=os.path.dirname(__file__) +'\databases\studentlogininfo.xlsx'
    else:
        loc=os.path.dirname(__file__) +'\databases\adminlogininfo.xlsx'
    check_file(loc)
    book =openpyxl.load_workbook(loc)
    sheet = book.active
    for i in range(1,sheet.max_row+1):
        if((sheet.cell(row=i,column=1)).value)==email:
            
            if((sheet.cell(row=i,column=2)).value)==password:
               
                return load(email)
            else:
                return 1
            
def add_xls_row(arr):
    
    '''
    Appends a row to a file
    '''
    workbook_name = os.path.dirname(__file__) +'\databases\studentinfo.xlsx'
    workbook2_name = os.path.dirname(__file__) +'\databases\studentlogininfo.xlsx'
    check_file(workbook_name,1)
    check_file(workbook2_name)

    wb =openpyxl.load_workbook(workbook_name)
    page = wb.active
    record = 10*[''] + [arr['username']] + 8*['']
    page.append(record)
    wb.save(filename =workbook_name)

    wb =openpyxl.load_workbook(workbook2_name)
    page = wb.active
    record = [arr['username'], arr['password']]
    page.append(record)
    wb.save(filename =workbook2_name)
    
def submit_save(value,row):
    
    '''
    overwrites information of particular row
    '''
    
    loc2=os.path.dirname(__file__) +'\databases\studentinfo.xlsx'
    check_file(loc2,1)
    book =openpyxl.load_workbook(loc2)
    sheet = book.active
    for i in range(1,20):
        sheet.cell(row=row,column=i).value = value[i-1]
        
    book.save(filename=loc2)
    
def registration_number():
    
    '''
    finds a new registration number
    '''
    
    loc2=os.path.dirname(__file__) +'\databases\studentinfo.xlsx'
    check_file(loc2,1)
    book =openpyxl.load_workbook(loc2)
    sheet = book.active
    
    lst = []
    for i in range(2,sheet.max_row+1):
        if sheet.cell(row=i,column=1).value != None:
            lst.append(int(sheet.cell(row=i,column=1).value[3:]))
    if len(lst) == 0:
        return '21/0'
    return '21/'+str(max(lst)+1)
    
if __name__ == '__main__':
    
    save_xls_file({'username':'abracadabra.@gmail.com', 'password':'124j/osk'},1)
    print(registration_number())